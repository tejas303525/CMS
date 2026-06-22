from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import uuid
import logging
import bcrypt
import jwt
from datetime import datetime, timezone, timedelta, date
from typing import List, Optional, Literal
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Query
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, ConfigDict

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# ---------- DB ----------
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_MINUTES = 60 * 8  # 8h for an admin tool

app = FastAPI(title="Church Management System")
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
log = logging.getLogger("cms")

# ---------- Helpers ----------
def now_utc() -> datetime:
    return datetime.now(timezone.utc)

def iso(dt: Optional[datetime]) -> Optional[str]:
    return dt.isoformat() if dt else None

def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False

def create_access_token(user_id: str, username: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "username": username,
        "role": role,
        "exp": now_utc() + timedelta(minutes=ACCESS_TOKEN_MINUTES),
        "type": "access",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

# ---------- Models ----------
Role = Literal["super_admin", "admin", "staff", "read_only"]
ROLE_LEVEL = {"read_only": 1, "staff": 2, "admin": 3, "super_admin": 4}

class UserPublic(BaseModel):
    id: str
    username: str
    full_name: str
    role: Role
    is_active: bool
    last_login: Optional[str] = None
    created_at: str

class LoginIn(BaseModel):
    username: str
    password: str

class UserCreateIn(BaseModel):
    username: str
    password: str
    full_name: str
    role: Role = "staff"

class UserUpdateIn(BaseModel):
    full_name: Optional[str] = None
    role: Optional[Role] = None
    is_active: Optional[bool] = None
    password: Optional[str] = None

class MemberIn(BaseModel):
    first_name: str
    middle_name: Optional[str] = ""
    last_name: str
    gender: Literal["Male", "Female", "Other"]
    date_of_birth: str  # ISO date
    membership_status: Literal["Active", "Inactive", "Visitor", "Transferred"] = "Active"
    membership_date: Optional[str] = None
    baptism_date: Optional[str] = None
    ministries: List[str] = []
    cell_group: Optional[str] = ""
    marital_status: Literal["Single", "Married", "Widowed", "Divorced"] = "Single"
    wedding_anniversary: Optional[str] = None
    occupation: Optional[str] = ""
    employer: Optional[str] = ""
    notes: Optional[str] = ""
    phone_primary: Optional[str] = ""
    phone_secondary: Optional[str] = ""
    whatsapp: Optional[str] = ""
    email: Optional[str] = ""
    address_street: Optional[str] = ""
    address_city: Optional[str] = ""
    country_origin: Optional[str] = ""
    country_current: Optional[str] = ""
    photo_url: Optional[str] = ""

class FamilyIn(BaseModel):
    family_name: str
    head_member_id: str
    members: List[dict] = []  # [{member_id, relationship_type}]

class ContributionIn(BaseModel):
    member_id: str
    contribution_date: str  # ISO date
    contribution_type: Literal["Tithe", "Offering", "Seed", "Pledge", "First Fruit", "Special"]
    amount: float
    payment_mode: Literal["Cash", "Bank Transfer", "Cheque", "Online", "POS"]
    reference_no: Optional[str] = ""
    notes: Optional[str] = ""

# ---------- Auth dependencies ----------
async def get_current_user(request: Request) -> dict:
    auth_header = request.headers.get("Authorization", "")
    token = auth_header[7:] if auth_header.startswith("Bearer ") else None
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user or not user.get("is_active", True):
            raise HTTPException(status_code=401, detail="User not found or inactive")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def require_role(min_role: Role):
    async def checker(user: dict = Depends(get_current_user)):
        if ROLE_LEVEL.get(user["role"], 0) < ROLE_LEVEL[min_role]:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return user
    return checker

# ---------- Audit ----------
async def audit(user: dict, action: str, entity: str, entity_id: str, details: Optional[dict] = None):
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "username": user["username"],
        "action": action,
        "entity": entity,
        "entity_id": entity_id,
        "details": details or {},
        "timestamp": now_utc().isoformat(),
    })

# ---------- Startup ----------
@app.on_event("startup")
async def startup():
    await db.users.create_index("username", unique=True)
    await db.members.create_index("member_id", unique=True)
    await db.contributions.create_index([("member_id", 1), ("contribution_date", -1)])
    await db.contributions.create_index("receipt_no", unique=True)
    await db.audit_logs.create_index([("timestamp", -1)])

    admin_username = os.environ.get("ADMIN_USERNAME", "superadmin")
    admin_password = os.environ.get("ADMIN_PASSWORD", "Admin@123")
    existing = await db.users.find_one({"username": admin_username})
    if not existing:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "username": admin_username,
            "password_hash": hash_password(admin_password),
            "full_name": "Super Administrator",
            "role": "super_admin",
            "is_active": True,
            "last_login": None,
            "created_at": now_utc().isoformat(),
        })
        log.info(f"Seeded super admin: {admin_username}")
    else:
        # keep password in sync with env on each boot (idempotent)
        if not verify_password(admin_password, existing["password_hash"]):
            await db.users.update_one({"username": admin_username}, {"$set": {"password_hash": hash_password(admin_password)}})

    # seed sample users (idempotent)
    for u in [
        {"username": "admin", "password": "Admin@123", "full_name": "Church Admin", "role": "admin"},
        {"username": "staff", "password": "Staff@123", "full_name": "Ministry Leader", "role": "staff"},
        {"username": "viewer", "password": "View@123", "full_name": "Cell Leader", "role": "read_only"},
    ]:
        if not await db.users.find_one({"username": u["username"]}):
            await db.users.insert_one({
                "id": str(uuid.uuid4()),
                "username": u["username"],
                "password_hash": hash_password(u["password"]),
                "full_name": u["full_name"],
                "role": u["role"],
                "is_active": True,
                "last_login": None,
                "created_at": now_utc().isoformat(),
            })


# ---------- Auth Endpoints ----------
@api.post("/auth/login")
async def login(body: LoginIn):
    user = await db.users.find_one({"username": body.username})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    if not user.get("is_active", True):
        raise HTTPException(status_code=403, detail="Account disabled")
    token = create_access_token(user["id"], user["username"], user["role"])
    await db.users.update_one({"id": user["id"]}, {"$set": {"last_login": now_utc().isoformat()}})
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": user["id"],
            "username": user["username"],
            "full_name": user["full_name"],
            "role": user["role"],
            "is_active": user["is_active"],
        }
    }

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user

# ---------- Users (Super Admin) ----------
@api.get("/users", response_model=List[UserPublic])
async def list_users(_: dict = Depends(require_role("super_admin"))):
    docs = await db.users.find({}, {"_id": 0, "password_hash": 0}).sort("created_at", -1).to_list(1000)
    return docs

@api.post("/users", response_model=UserPublic)
async def create_user(body: UserCreateIn, user: dict = Depends(require_role("super_admin"))):
    if await db.users.find_one({"username": body.username}):
        raise HTTPException(status_code=400, detail="Username already exists")
    doc = {
        "id": str(uuid.uuid4()),
        "username": body.username,
        "password_hash": hash_password(body.password),
        "full_name": body.full_name,
        "role": body.role,
        "is_active": True,
        "last_login": None,
        "created_at": now_utc().isoformat(),
    }
    await db.users.insert_one(doc)
    await audit(user, "create", "user", doc["id"], {"username": body.username, "role": body.role})
    doc.pop("password_hash", None)
    return doc

@api.patch("/users/{user_id}", response_model=UserPublic)
async def update_user(user_id: str, body: UserUpdateIn, user: dict = Depends(require_role("super_admin"))):
    update = {}
    if body.full_name is not None: update["full_name"] = body.full_name
    if body.role is not None: update["role"] = body.role
    if body.is_active is not None: update["is_active"] = body.is_active
    if body.password: update["password_hash"] = hash_password(body.password)
    if update:
        await db.users.update_one({"id": user_id}, {"$set": update})
    doc = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="User not found")
    await audit(user, "update", "user", user_id, {"fields": list(update.keys())})
    return doc

@api.delete("/users/{user_id}")
async def delete_user(user_id: str, user: dict = Depends(require_role("super_admin"))):
    if user_id == user["id"]:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    await db.users.delete_one({"id": user_id})
    await audit(user, "delete", "user", user_id)
    return {"ok": True}

# ---------- Members ----------
async def _next_member_id() -> str:
    # Format CHM00001
    last = await db.counters.find_one_and_update(
        {"_id": "member_id"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    seq = last["seq"] if last else 1
    return f"CHM{seq:05d}"

@api.get("/members")
async def list_members(
    q: Optional[str] = None,
    status: Optional[str] = None,
    ministry: Optional[str] = None,
    birthday_month: Optional[int] = None,
    anniversary_month: Optional[int] = None,
    limit: int = 200,
    user: dict = Depends(get_current_user),
):
    query = {}
    if status: query["membership_status"] = status
    if ministry: query["ministries"] = ministry
    if q:
        rx = {"$regex": q, "$options": "i"}
        query["$or"] = [
            {"first_name": rx}, {"last_name": rx}, {"middle_name": rx},
            {"member_id": rx}, {"phone_primary": rx}, {"phone_secondary": rx},
            {"whatsapp": rx}, {"email": rx},
        ]
    docs = await db.members.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    # client-side filter for month based fields
    def month_match(value: Optional[str], m: int) -> bool:
        if not value:
            return False
        try:
            return int(value[5:7]) == m
        except Exception:
            return False
    if birthday_month:
        docs = [d for d in docs if month_match(d.get("date_of_birth"), birthday_month)]
    if anniversary_month:
        docs = [d for d in docs if month_match(d.get("wedding_anniversary"), anniversary_month)]
    return docs

@api.post("/members")
async def create_member(body: MemberIn, user: dict = Depends(require_role("staff"))):
    mid = await _next_member_id()
    doc = body.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["member_id"] = mid
    doc["created_at"] = now_utc().isoformat()
    doc["updated_at"] = doc["created_at"]
    doc["created_by"] = user["username"]
    await db.members.insert_one(doc)
    await audit(user, "create", "member", doc["id"], {"member_id": mid})
    doc.pop("_id", None)
    return doc

@api.get("/members/{member_id}")
async def get_member(member_id: str, user: dict = Depends(get_current_user)):
    doc = await db.members.find_one({"id": member_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Member not found")
    # strip pastoral notes for non-admin
    if ROLE_LEVEL.get(user["role"], 0) < ROLE_LEVEL["admin"]:
        doc["notes"] = ""
    return doc

@api.patch("/members/{member_id}")
async def update_member(member_id: str, body: MemberIn, user: dict = Depends(require_role("staff"))):
    update = body.model_dump()
    update["updated_at"] = now_utc().isoformat()
    res = await db.members.update_one({"id": member_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Member not found")
    await audit(user, "update", "member", member_id)
    doc = await db.members.find_one({"id": member_id}, {"_id": 0})
    return doc

@api.delete("/members/{member_id}")
async def soft_delete_member(member_id: str, user: dict = Depends(require_role("admin"))):
    res = await db.members.update_one({"id": member_id}, {"$set": {"membership_status": "Inactive", "updated_at": now_utc().isoformat()}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Member not found")
    await audit(user, "deactivate", "member", member_id)
    return {"ok": True}

# ---------- Families ----------
@api.get("/families")
async def list_families(user: dict = Depends(get_current_user)):
    docs = await db.families.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return docs

@api.post("/families")
async def create_family(body: FamilyIn, user: dict = Depends(require_role("staff"))):
    doc = body.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = now_utc().isoformat()
    doc["created_by"] = user["username"]
    await db.families.insert_one(doc)
    await audit(user, "create", "family", doc["id"])
    doc.pop("_id", None)
    return doc

@api.get("/families/{family_id}")
async def get_family(family_id: str, user: dict = Depends(get_current_user)):
    fam = await db.families.find_one({"id": family_id}, {"_id": 0})
    if not fam:
        raise HTTPException(status_code=404, detail="Family not found")
    # enrich with member info
    ids = [fam["head_member_id"]] + [m["member_id"] for m in fam.get("members", [])]
    members = await db.members.find({"id": {"$in": ids}}, {"_id": 0}).to_list(100)
    members_by_id = {m["id"]: m for m in members}
    fam["head_member"] = members_by_id.get(fam["head_member_id"])
    fam["enriched_members"] = []
    for m in fam.get("members", []):
        em = members_by_id.get(m["member_id"])
        if em:
            fam["enriched_members"].append({**em, "relationship_type": m.get("relationship_type", "Other")})
    return fam

@api.patch("/families/{family_id}")
async def update_family(family_id: str, body: FamilyIn, user: dict = Depends(require_role("staff"))):
    update = body.model_dump()
    update["updated_at"] = now_utc().isoformat()
    res = await db.families.update_one({"id": family_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Family not found")
    await audit(user, "update", "family", family_id)
    doc = await db.families.find_one({"id": family_id}, {"_id": 0})
    return doc

@api.delete("/families/{family_id}")
async def delete_family(family_id: str, user: dict = Depends(require_role("admin"))):
    await db.families.delete_one({"id": family_id})
    await audit(user, "delete", "family", family_id)
    return {"ok": True}

# ---------- Contributions ----------
async def _next_receipt_no() -> str:
    last = await db.counters.find_one_and_update(
        {"_id": "receipt_no"}, {"$inc": {"seq": 1}}, upsert=True, return_document=True
    )
    seq = last["seq"] if last else 1
    return f"RCP{seq:06d}"

@api.get("/contributions")
async def list_contributions(
    member_id: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    type: Optional[str] = None,
    user: dict = Depends(require_role("admin")),
):
    query = {}
    if member_id: query["member_id"] = member_id
    if year: query["year"] = year
    if month: query["month"] = month
    if type: query["contribution_type"] = type
    docs = await db.contributions.find(query, {"_id": 0}).sort("contribution_date", -1).to_list(1000)
    return docs

@api.post("/contributions")
async def create_contribution(body: ContributionIn, user: dict = Depends(require_role("admin"))):
    member = await db.members.find_one({"id": body.member_id}, {"_id": 0, "first_name": 1, "last_name": 1, "member_id": 1, "id": 1})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    dt = datetime.fromisoformat(body.contribution_date)
    doc = body.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["receipt_no"] = await _next_receipt_no()
    doc["currency"] = "INR"
    doc["year"] = dt.year
    doc["month"] = dt.month
    doc["member_name"] = f"{member['first_name']} {member['last_name']}"
    doc["member_external_id"] = member["member_id"]
    doc["recorded_by"] = user["username"]
    doc["created_at"] = now_utc().isoformat()
    await db.contributions.insert_one(doc)
    await audit(user, "create", "contribution", doc["id"], {"receipt_no": doc["receipt_no"], "amount": doc["amount"]})
    doc.pop("_id", None)
    return doc

@api.delete("/contributions/{contribution_id}")
async def delete_contribution(contribution_id: str, user: dict = Depends(require_role("admin"))):
    await db.contributions.delete_one({"id": contribution_id})
    await audit(user, "delete", "contribution", contribution_id)
    return {"ok": True}

@api.get("/contributions/{contribution_id}/receipt")
async def contribution_receipt(contribution_id: str, user: dict = Depends(require_role("admin"))):
    c = await db.contributions.find_one({"id": contribution_id}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Contribution not found")
    member = await db.members.find_one({"id": c["member_id"]}, {"_id": 0}) or {}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=40, bottomMargin=40, leftMargin=48, rightMargin=48)
    styles = getSampleStyleSheet()
    brand = colors.HexColor("#2A4B3C")
    accent = colors.HexColor("#C49A45")
    muted = colors.HexColor("#57534E")

    title_style = ParagraphStyle("Title", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=22, textColor=brand, alignment=0)
    label_style = ParagraphStyle("Lbl", parent=styles["Normal"], fontSize=8, textColor=muted, leading=10, spaceAfter=2)
    value_style = ParagraphStyle("Val", parent=styles["Normal"], fontSize=11, textColor=colors.HexColor("#1C1917"), leading=14)
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=9, textColor=muted)
    big_amount = ParagraphStyle("Amt", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=28, textColor=brand, alignment=0)

    elements = []
    # Header band
    header_tbl = Table([[
        Paragraph("OFFICIAL RECEIPT", ParagraphStyle("H", parent=styles["Normal"], fontSize=9, textColor=accent, leading=12, spaceAfter=4)),
    ], [
        Paragraph("Church Management System", title_style)
    ]], colWidths=[None])
    header_tbl.setStyle(TableStyle([("BOTTOMPADDING", (0, 0), (-1, -1), 0), ("TOPPADDING", (0, 0), (-1, -1), 0)]))
    elements += [header_tbl, Spacer(1, 6),
                 Paragraph(f"Receipt No: <b>{c['receipt_no']}</b> &nbsp;&nbsp;|&nbsp;&nbsp; Date: <b>{c['contribution_date']}</b>", small),
                 Spacer(1, 18)]

    # Member block
    member_name = f"{member.get('first_name','')} {member.get('last_name','')}".strip() or c.get("member_name", "")
    member_tbl = Table([
        [Paragraph("RECEIVED FROM", label_style)],
        [Paragraph(member_name, value_style)],
        [Paragraph(f"Member ID: {member.get('member_id', c.get('member_external_id',''))}", small)],
    ], colWidths=[None])
    member_tbl.setStyle(TableStyle([("BOTTOMPADDING", (0, 0), (-1, -1), 1), ("TOPPADDING", (0, 0), (-1, -1), 1)]))
    elements += [member_tbl, Spacer(1, 14)]

    # Amount block (highlighted)
    amount_box = Table([
        [Paragraph("AMOUNT RECEIVED", label_style)],
        [Paragraph(f"INR {c['amount']:,.2f}", big_amount)],
    ], colWidths=[None])
    amount_box.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F2EA")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E8E4D9")),
        ("LEFTPADDING", (0, 0), (-1, -1), 16),
        ("RIGHTPADDING", (0, 0), (-1, -1), 16),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    elements += [amount_box, Spacer(1, 18)]

    # Details grid
    rows = [
        ["Contribution Type", c.get("contribution_type", "")],
        ["Payment Mode", c.get("payment_mode", "")],
        ["Reference No.", c.get("reference_no") or "—"],
        ["Currency", "INR"],
        ["Recorded By", c.get("recorded_by", "")],
        ["Notes", c.get("notes") or "—"],
    ]
    details = Table(rows, colWidths=[140, None])
    details.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (0, -1), muted),
        ("TEXTCOLOR", (1, 0), (1, -1), colors.HexColor("#1C1917")),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("LINEBELOW", (0, 0), (-1, -2), 0.25, colors.HexColor("#E8E4D9")),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements += [details, Spacer(1, 24)]

    # Footer
    elements += [
        Paragraph("Thank you for your faithful giving.", ParagraphStyle("F1", parent=styles["Normal"], fontSize=11, textColor=brand, fontName="Helvetica-Oblique")),
        Spacer(1, 24),
        Table([["", ""]], colWidths=[180, 180], style=TableStyle([
            ("LINEABOVE", (0, 0), (0, 0), 0.5, colors.HexColor("#1C1917")),
            ("LINEABOVE", (1, 0), (1, 0), 0.5, colors.HexColor("#1C1917")),
        ])),
        Table([
            [Paragraph("Authorized signature", small), Paragraph("Recipient signature", small)],
        ], colWidths=[180, 180]),
        Spacer(1, 30),
        Paragraph(f"Generated {now_utc().strftime('%Y-%m-%d %H:%M UTC')} · This is a system-generated receipt.", small),
    ]

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="receipt_{c["receipt_no"]}.pdf"'},
    )

@api.get("/contributions/summary/{member_id}")
async def member_contribution_summary(member_id: str, year: int, user: dict = Depends(require_role("admin"))):
    docs = await db.contributions.find({"member_id": member_id, "year": year}, {"_id": 0}).to_list(1000)
    months = {m: {"Tithe": 0.0, "Offering": 0.0, "Other": 0.0, "Total": 0.0} for m in range(1, 13)}
    for d in docs:
        m = d["month"]
        bucket = "Tithe" if d["contribution_type"] == "Tithe" else ("Offering" if d["contribution_type"] == "Offering" else "Other")
        months[m][bucket] += d["amount"]
        months[m]["Total"] += d["amount"]
    annual_total = sum(v["Total"] for v in months.values())
    return {"year": year, "months": months, "annual_total": annual_total, "transactions": docs}

# ---------- Dashboard ----------
@api.get("/dashboard/summary")
async def dashboard_summary(user: dict = Depends(get_current_user)):
    today = now_utc().date()
    month_start = today.replace(day=1)
    last_month_start = (month_start - timedelta(days=1)).replace(day=1)

    active_members = await db.members.count_documents({"membership_status": "Active"})
    new_members_this_month = await db.members.count_documents({"created_at": {"$gte": month_start.isoformat()}})

    # tithes only for admin+
    show_finance = ROLE_LEVEL.get(user["role"], 0) >= ROLE_LEVEL["admin"]
    this_month_total = 0.0
    last_month_total = 0.0
    if show_finance:
        cur = db.contributions.find({"year": today.year, "month": today.month})
        async for d in cur:
            this_month_total += d.get("amount", 0)
        lm_year = last_month_start.year
        lm_month = last_month_start.month
        cur = db.contributions.find({"year": lm_year, "month": lm_month})
        async for d in cur:
            last_month_total += d.get("amount", 0)

    # upcoming birthdays / anniversaries (next 7 days, by month-day)
    members = await db.members.find({"membership_status": {"$ne": "Inactive"}}, {"_id": 0}).to_list(5000)

    def upcoming(field: str, days: int = 7):
        items = []
        for m in members:
            v = m.get(field)
            if not v: continue
            try:
                d = datetime.fromisoformat(v).date()
            except Exception:
                continue
            this_year = d.replace(year=today.year)
            if this_year < today:
                this_year = d.replace(year=today.year + 1)
            delta = (this_year - today).days
            if 0 <= delta <= days:
                items.append({
                    "id": m["id"],
                    "member_id": m.get("member_id"),
                    "name": f"{m['first_name']} {m['last_name']}",
                    "date": v,
                    "next_occurrence": this_year.isoformat(),
                    "in_days": delta,
                })
        items.sort(key=lambda x: x["in_days"])
        return items

    upcoming_birthdays = upcoming("date_of_birth")
    upcoming_anniversaries = upcoming("wedding_anniversary")

    # recent activity
    recent = await db.audit_logs.find({}, {"_id": 0}).sort("timestamp", -1).to_list(10)

    return {
        "active_members": active_members,
        "new_members_this_month": new_members_this_month,
        "tithes_this_month": this_month_total,
        "tithes_last_month": last_month_total,
        "upcoming_birthdays": upcoming_birthdays,
        "upcoming_anniversaries": upcoming_anniversaries,
        "recent_activity": recent,
        "show_finance": show_finance,
    }

# ---------- Audit ----------
@api.get("/audit-logs")
async def list_audit(limit: int = 100, user: dict = Depends(require_role("admin"))):
    docs = await db.audit_logs.find({}, {"_id": 0}).sort("timestamp", -1).to_list(limit)
    return docs

# ---------- Reports ----------
def _xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

def _pdf_response(title: str, headers: List[str], rows: List[List[str]], filename: str, subtitle: str = "") -> StreamingResponse:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=18, textColor=colors.HexColor("#2A4B3C"))
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor("#57534E"))
    elements = [Paragraph(title, title_style), Spacer(1, 6)]
    if subtitle:
        elements += [Paragraph(subtitle, sub_style), Spacer(1, 6)]
    elements += [Paragraph(f"Generated: {now_utc().strftime('%Y-%m-%d %H:%M UTC')}", sub_style), Spacer(1, 12)]
    table_data = [headers] + rows
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2A4B3C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E8E4D9")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FDFBF7")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})

async def _members_dataset(birthday_month: Optional[int] = None, anniversary_month: Optional[int] = None, status: Optional[str] = None):
    q = {}
    if status: q["membership_status"] = status
    docs = await db.members.find(q, {"_id": 0}).to_list(5000)
    def month_match(value, m):
        if not value: return False
        try: return int(value[5:7]) == m
        except: return False
    if birthday_month:
        docs = [d for d in docs if month_match(d.get("date_of_birth"), birthday_month)]
    if anniversary_month:
        docs = [d for d in docs if month_match(d.get("wedding_anniversary"), anniversary_month)]
    return docs

@api.get("/reports/members")
async def report_members(format: Literal["pdf", "excel"] = "excel", status: Optional[str] = None, user: dict = Depends(require_role("staff"))):
    docs = await _members_dataset(status=status)
    headers = ["Member ID", "Full Name", "Gender", "DOB", "Status", "Phone", "Email", "City"]
    rows = [[d.get("member_id",""), f"{d.get('first_name','')} {d.get('last_name','')}", d.get("gender",""), d.get("date_of_birth",""), d.get("membership_status",""), d.get("phone_primary",""), d.get("email",""), d.get("address_city","")] for d in docs]
    title = "Member Directory"
    if format == "excel":
        wb = Workbook(); ws = wb.active; ws.title = "Members"
        ws.append(headers)
        for r in rows: ws.append(r)
        return _xlsx_response(wb, "member_directory.xlsx")
    return _pdf_response(title, headers, rows, "member_directory.pdf")

@api.get("/reports/birthdays")
async def report_birthdays(month: int = Query(..., ge=1, le=12), format: Literal["pdf", "excel"] = "excel", user: dict = Depends(require_role("staff"))):
    docs = await _members_dataset(birthday_month=month)
    headers = ["Member ID", "Full Name", "DOB", "Phone", "Email"]
    rows = [[d.get("member_id",""), f"{d.get('first_name','')} {d.get('last_name','')}", d.get("date_of_birth",""), d.get("phone_primary",""), d.get("email","")] for d in docs]
    month_name = datetime(2000, month, 1).strftime("%B")
    if format == "excel":
        wb = Workbook(); ws = wb.active; ws.title = "Birthdays"
        ws.append(headers)
        for r in rows: ws.append(r)
        return _xlsx_response(wb, f"birthdays_{month_name}.xlsx")
    return _pdf_response(f"Birthdays — {month_name}", headers, rows, f"birthdays_{month_name}.pdf")

@api.get("/reports/anniversaries")
async def report_anniversaries(month: int = Query(..., ge=1, le=12), format: Literal["pdf", "excel"] = "excel", user: dict = Depends(require_role("staff"))):
    docs = await _members_dataset(anniversary_month=month)
    headers = ["Member ID", "Full Name", "Anniversary", "Phone", "Email"]
    rows = [[d.get("member_id",""), f"{d.get('first_name','')} {d.get('last_name','')}", d.get("wedding_anniversary",""), d.get("phone_primary",""), d.get("email","")] for d in docs]
    month_name = datetime(2000, month, 1).strftime("%B")
    if format == "excel":
        wb = Workbook(); ws = wb.active; ws.title = "Anniversaries"
        ws.append(headers)
        for r in rows: ws.append(r)
        return _xlsx_response(wb, f"anniversaries_{month_name}.xlsx")
    return _pdf_response(f"Anniversaries — {month_name}", headers, rows, f"anniversaries_{month_name}.pdf")

@api.get("/reports/contributions-monthly")
async def report_contrib_monthly(year: int, month: int = Query(..., ge=1, le=12), format: Literal["pdf", "excel"] = "excel", user: dict = Depends(require_role("admin"))):
    docs = await db.contributions.find({"year": year, "month": month}, {"_id": 0}).to_list(5000)
    headers = ["Date", "Receipt #", "Member ID", "Member", "Type", "Amount (INR)", "Payment Mode", "Reference"]
    rows = [[d.get("contribution_date",""), d.get("receipt_no",""), d.get("member_external_id",""), d.get("member_name",""), d.get("contribution_type",""), f"{d.get('amount',0):.2f}", d.get("payment_mode",""), d.get("reference_no","")] for d in docs]
    total = sum(d.get("amount", 0) for d in docs)
    rows.append(["", "", "", "", "TOTAL", f"{total:.2f}", "", ""])
    month_name = datetime(2000, month, 1).strftime("%B")
    title = f"Monthly Contributions — {month_name} {year}"
    if format == "excel":
        wb = Workbook(); ws = wb.active; ws.title = "Contributions"
        ws.append(headers)
        for r in rows: ws.append(r)
        return _xlsx_response(wb, f"contributions_{year}_{month:02d}.xlsx")
    return _pdf_response(title, headers, rows, f"contributions_{year}_{month:02d}.pdf")

@api.get("/reports/member-statement/{member_id}")
async def report_member_statement(member_id: str, year: int, format: Literal["pdf", "excel"] = "pdf", user: dict = Depends(require_role("admin"))):
    member = await db.members.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    summary = await member_contribution_summary(member_id, year, user)  # type: ignore
    headers = ["Month", "Tithe (INR)", "Offering (INR)", "Other (INR)", "Total (INR)"]
    rows = []
    for m in range(1, 13):
        v = summary["months"][m]
        rows.append([datetime(2000, m, 1).strftime("%B"), f"{v['Tithe']:.2f}", f"{v['Offering']:.2f}", f"{v['Other']:.2f}", f"{v['Total']:.2f}"])
    rows.append(["ANNUAL TOTAL", "", "", "", f"{summary['annual_total']:.2f}"])
    name = f"{member['first_name']} {member['last_name']}"
    title = f"Contribution Statement — {name} ({year})"
    if format == "excel":
        wb = Workbook(); ws = wb.active; ws.title = "Statement"
        ws.append(headers)
        for r in rows: ws.append(r)
        return _xlsx_response(wb, f"statement_{member['member_id']}_{year}.xlsx")
    return _pdf_response(title, headers, rows, f"statement_{member['member_id']}_{year}.pdf", subtitle=f"Member ID: {member['member_id']}")

@api.get("/reports/non-contributing")
async def report_non_contrib(months: int = 2, format: Literal["pdf", "excel"] = "excel", user: dict = Depends(require_role("admin"))):
    today = now_utc().date()
    cutoff = today - timedelta(days=months * 30)
    cutoff_iso = cutoff.isoformat()
    members = await db.members.find({"membership_status": "Active"}, {"_id": 0}).to_list(5000)
    rows_out = []
    for m in members:
        last_c = await db.contributions.find_one({"member_id": m["id"]}, sort=[("contribution_date", -1)])
        last_date = last_c["contribution_date"] if last_c else None
        if not last_date or last_date < cutoff_iso:
            rows_out.append([m.get("member_id",""), f"{m.get('first_name','')} {m.get('last_name','')}", m.get("phone_primary",""), m.get("email",""), last_date or "—"])
    headers = ["Member ID", "Full Name", "Phone", "Email", "Last Contribution"]
    title = f"Non-Contributing Members (no record in {months} months)"
    if format == "excel":
        wb = Workbook(); ws = wb.active; ws.title = "Non-Contributing"
        ws.append(headers)
        for r in rows_out: ws.append(r)
        return _xlsx_response(wb, "non_contributing.xlsx")
    return _pdf_response(title, headers, rows_out, "non_contributing.pdf")

@api.get("/reports/families")
async def report_families(format: Literal["pdf", "excel"] = "excel", user: dict = Depends(require_role("staff"))):
    fams = await db.families.find({}, {"_id": 0}).to_list(2000)
    headers = ["Family Name", "Head", "Member Count"]
    rows = []
    for f in fams:
        head = await db.members.find_one({"id": f["head_member_id"]}, {"_id": 0, "first_name": 1, "last_name": 1})
        head_name = f"{head['first_name']} {head['last_name']}" if head else "—"
        rows.append([f["family_name"], head_name, len(f.get("members", []))])
    title = "Family Report"
    if format == "excel":
        wb = Workbook(); ws = wb.active; ws.title = "Families"
        ws.append(headers)
        for r in rows: ws.append(r)
        return _xlsx_response(wb, "families.xlsx")
    return _pdf_response(title, headers, rows, "families.pdf")

# ---------- Health ----------
@api.get("/")
async def root():
    return {"app": "Church Management System", "version": "1.0"}

app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown():
    client.close()



